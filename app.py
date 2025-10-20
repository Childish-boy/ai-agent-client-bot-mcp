from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from langchain.agents import AgentExecutor, create_openai_tools_agent
from langchain.tools import tool
from langchain_openai import ChatOpenAI
from langchain.prompts import ChatPromptTemplate, MessagesPlaceholder
import requests
import json
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import base64
import uuid
from pathlib import Path
from mcp_client import get_mcp_client

# 加载环境变量
load_dotenv()

# 加载画图配置
def load_image_config():
    """加载千问AI画图配置"""
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"加载配置文件失败: {e}")
        return None

IMAGE_CONFIG = load_image_config()

# 创建图片保存目录
if IMAGE_CONFIG:
    IMAGES_DIR = IMAGE_CONFIG.get('save_directory', './generated_images')
    Path(IMAGES_DIR).mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
CORS(app)

# 模拟员工日程数据
def generate_schedule_data(department=None):
    """生成模拟的员工日程数据"""
    departments = {
        "技术部": [
            {"name": "张三", "position": "高级工程师", "tasks": ["开发新功能模块", "代码审查", "技术文档编写"]},
            {"name": "李四", "position": "前端工程师", "tasks": ["UI界面优化", "响应式布局调整", "前端性能优化"]},
            {"name": "王五", "position": "后端工程师", "tasks": ["API接口开发", "数据库优化", "服务器维护"]},
        ],
        "市场部": [
            {"name": "赵六", "position": "市场经理", "tasks": ["市场调研", "营销方案策划", "客户拜访"]},
            {"name": "钱七", "position": "市场专员", "tasks": ["社交媒体运营", "活动策划执行", "数据分析报告"]},
        ],
        "人事部": [
            {"name": "孙八", "position": "人事经理", "tasks": ["招聘面试", "员工培训", "绩效考核"]},
            {"name": "周九", "position": "人事专员", "tasks": ["员工档案管理", "考勤统计", "福利发放"]},
        ],
        "财务部": [
            {"name": "吴十", "position": "财务经理", "tasks": ["财务报表审核", "预算编制", "税务申报"]},
            {"name": "郑十一", "position": "会计", "tasks": ["日常账务处理", "发票管理", "费用报销审核"]},
        ]
    }
    
    today = datetime.now()
    schedule_data = []
    
    # 根据部门筛选
    if department and department != "全体":
        if department in departments:
            target_deps = {department: departments[department]}
        else:
            return {"error": f"未找到部门: {department}", "available_departments": list(departments.keys())}
    else:
        target_deps = departments
    
    for dept_name, employees in target_deps.items():
        for emp in employees:
            for i, task in enumerate(emp["tasks"]):
                schedule_date = today + timedelta(days=i)
                schedule_data.append({
                    "department": dept_name,
                    "employee_name": emp["name"],
                    "position": emp["position"],
                    "date": schedule_date.strftime("%Y-%m-%d"),
                    "task": task,
                    "status": "进行中" if i == 0 else "待开始",
                    "priority": "高" if i == 0 else "中"
                })
    
    return {
        "success": True,
        "total_count": len(schedule_data),
        "query_department": department if department else "全体",
        "schedules": schedule_data
    }


# Flask 接口：获取员工日程表
@app.route('/api/employee-schedule', methods=['POST'])
def get_employee_schedule():
    """
    获取员工日程表的 API 接口
    请求体: {"department": "技术部"} 或 {"department": "全体"}
    """
    data = request.get_json()
    department = data.get('department', '全体') if data else '全体'
    
    schedule_data = generate_schedule_data(department)
    return jsonify(schedule_data)


# LangChain 工具定义
@tool
def get_company_schedule(department: str = "全体") -> str:
    """
    获取公司员工日程安排的工具。
    
    参数:
        department: 部门名称，可选值为 "全体"、"技术部"、"市场部"、"人事部"、"财务部"
    
    返回:
        JSON格式的员工日程数据
    """
    try:
        # 调用内部接口
        response = requests.post(
            'http://localhost:5000/api/employee-schedule',
            json={"department": department},
            timeout=10
        )
        
        if response.status_code == 200:
            data = response.json()
            return json.dumps(data, ensure_ascii=False, indent=2)
        else:
            return f"获取日程失败，状态码: {response.status_code}"
    except Exception as e:
        return f"调用接口出错: {str(e)}"


@tool
def generate_image(prompt: str) -> str:
    """
    使用千问AI生成图片的工具。
    
    参数:
        prompt: 图片描述文本，例如："一只可爱的小猫在花园里玩耍"
    
    返回:
        JSON格式的图片信息，包含图片ID和URL
    """
    if not IMAGE_CONFIG:
        return json.dumps({"error": "画图配置未加载"}, ensure_ascii=False)
    
    try:
        # 千问AI画图API配置
        api_key = IMAGE_CONFIG['tongyi']['api_key']
        endpoint = IMAGE_CONFIG['tongyi']['endpoint']
        model = IMAGE_CONFIG['tongyi']['model']
        
        # 请求参数
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {api_key}',
            'X-DashScope-Async': 'enable'  # 异步调用
        }
        
        params = {
            'model': model,
            'input': {
                'prompt': prompt
            },
            'parameters': {
                'size': f"{IMAGE_CONFIG['default_parameters']['width']}*{IMAGE_CONFIG['default_parameters']['height']}",
                'n': IMAGE_CONFIG['default_parameters']['num_images']
            }
        }
        
        # 提交任务
        response = requests.post(endpoint, headers=headers, json=params, timeout=30)
        result = response.json()
        
        if response.status_code != 200:
            return json.dumps({"error": f"API调用失败: {result}"}, ensure_ascii=False)
        
        # 获取任务ID
        task_id = result.get('output', {}).get('task_id')
        if not task_id:
            return json.dumps({"error": "未获取到任务ID"}, ensure_ascii=False)
        
        # 查询任务结果（轮询）
        query_url = f"https://dashscope.aliyuncs.com/api/v1/tasks/{task_id}"
        max_retries = 30
        
        for i in range(max_retries):
            import time
            time.sleep(2)  # 等待2秒
            
            query_response = requests.get(query_url, headers={'Authorization': f'Bearer {api_key}'})
            query_result = query_response.json()
            
            task_status = query_result.get('output', {}).get('task_status')
            
            if task_status == 'SUCCEEDED':
                # 获取图片URL
                image_url = query_result.get('output', {}).get('results', [{}])[0].get('url')
                
                if not image_url:
                    return json.dumps({"error": "未获取到图片URL"}, ensure_ascii=False)
                
                # 下载图片
                image_response = requests.get(image_url, timeout=30)
                if image_response.status_code != 200:
                    return json.dumps({"error": "图片下载失败"}, ensure_ascii=False)
                
                # 保存图片
                image_id = str(uuid.uuid4())
                image_filename = f"{image_id}.{IMAGE_CONFIG['save_format']}"
                image_path = os.path.join(IMAGES_DIR, image_filename)
                
                with open(image_path, 'wb') as f:
                    f.write(image_response.content)
                
                # 返回图片信息
                return json.dumps({
                    "success": True,
                    "image_id": image_id,
                    "filename": image_filename,
                    "url": f"/api/images/{image_filename}",
                    "prompt": prompt
                }, ensure_ascii=False)
            
            elif task_status == 'FAILED':
                return json.dumps({"error": "图片生成失败"}, ensure_ascii=False)
        
        return json.dumps({"error": "图片生成超时"}, ensure_ascii=False)
        
    except Exception as e:
        return json.dumps({"error": f"生成图片时出错: {str(e)}"}, ensure_ascii=False)


@tool
def query_weather(city: str, forecast: bool = False, filter_indices: str = None) -> str:
    """
    查询指定城市的天气信息（通过 MCP HTTP 服务）
    
    ⭐ 工作流程：
    1. MCP 服务返回完整的天气数据（所有可用天数）
    2. 你理解用户需求，决定需要哪些天的数据
    3. 通过 filter_indices 参数告诉工具要保留哪些天（用于前端展示）
    4. 在文本回答中，你根据用户需求进行详细说明
    
    参数:
        city: 城市名称，如：北京、上海、广州
        forecast: 是否查询未来天气预报
                 - False（默认）：返回实时天气
                 - True：返回未来天气预报
        filter_indices: 可选，用于前端格式化展示时筛选数据
                 - 格式：用逗号分隔的索引，如 "0,1,2" 表示前3天，"2,3" 表示第3、4天
                 - 索引从0开始：0=第1天(明天), 1=第2天(后天), 2=第3天, 3=第4天
                 - 示例：
                   * 用户说"未来3天" → filter_indices="0,1,2"
                   * 用户说"明天和后天" → filter_indices="0,1"
                   * 用户说"周三和周四"（假设是第3、4天）→ filter_indices="2,3"
                   * 用户说"后天" → filter_indices="1"
                 - 如果不传或传 None，返回所有数据
    
    返回:
        JSON格式的天气信息
        - 实时天气: {type: "current", city, weather, temperature, ...}
        - 预报天气: {type: "forecast", city, forecasts: [...]}（根据 filter_indices 筛选后的）
    
    ⚠️ 重要：
      - filter_indices 只影响返回的结构化数据（用于前端卡片展示）
      - 你的文本回答应该与 filter_indices 保持一致
      - 这样前端展示的卡片就会和你的文字说明一致
    """
    weather_type = "预报天气" if forecast else "实时天气"
    filter_info = f"（筛选: {filter_indices}）" if filter_indices else "（完整数据）"
    print(f"🔄 通过 MCP 协议查询{weather_type}{filter_info}: {city}")
    
    import json
    
    # 获取 MCP 客户端（连接到远程 FastMCP 服务器）
    client = get_mcp_client("http://localhost:8001")
    
    # 根据类型调用不同的 MCP 工具
    if forecast:
        result = client.query_weather_forecast(city)
    else:
        result = client.query_current_weather(city)
    
    # 如果是预报天气且指定了筛选索引，进行筛选
    if forecast and filter_indices and result.get('success') and result.get('type') == 'forecast':
        try:
            indices = [int(idx.strip()) for idx in filter_indices.split(',')]
            original_forecasts = result.get('forecasts', [])
            filtered_forecasts = [original_forecasts[i] for i in indices if i < len(original_forecasts)]
            result['forecasts'] = filtered_forecasts
            result['filtered'] = True
            print(f"✅ 筛选后返回 {len(filtered_forecasts)} 天数据（索引: {indices}）")
        except Exception as e:
            print(f"⚠️ 筛选失败: {e}，返回完整数据")
    
    return json.dumps(result, ensure_ascii=False)


# 初始化 DeepSeek LLM
def init_llm():
    """初始化 DeepSeek 大模型"""
    api_key = os.getenv('DEEPSEEK_API_KEY')
    
    if not api_key:
        raise ValueError("请设置 DEEPSEEK_API_KEY 环境变量")
    
    # DeepSeek 使用 OpenAI 兼容接口
    llm = ChatOpenAI(
        model="deepseek-chat",
        openai_api_key=api_key,
        openai_api_base="https://api.deepseek.com/v1",
        temperature=0.7,
    )
    
    return llm


# 创建 Agent
def create_agent():
    """创建 LangChain Agent"""
    llm = init_llm()
    
    # 定义工具列表
    tools = [get_company_schedule, generate_image, query_weather]
    
    # 创建 Prompt
    prompt = ChatPromptTemplate.from_messages([
        ("system", """你是一个专业的AI助手，可以进行普通对话，也具备特殊功能。

💬 **普通对话能力**：
- 可以回答各种问题：编程、知识问答、写作等
- 可以帮助用户编写代码、解决问题
- 可以进行友好的日常对话
- 如果问题不需要使用特殊工具，就直接回答

🛠️ **特殊工具功能**：

📋 日程管理 - 使用 get_company_schedule 工具：
  - 当用户询问"员工日程"、"工作安排"时使用
  - 可用部门：技术部、市场部、人事部、财务部、全体
  - 使用后会返回JSON数据

🎨 AI画图 - 使用 generate_image 工具：
  - 当用户说"画"、"生成图片"、"创作"时使用
  - 支持各种风格：写实、卡通、艺术等
  - 使用后会返回JSON格式的图片信息
  - ⚠️ 重要：当你调用 generate_image 工具后，必须简短回复，不要重复图片信息

🌤️ 天气查询 - 使用 query_weather 工具：
  - 当用户询问"天气"、"气温"、"天气怎么样"时使用
  - 支持全国各大城市查询
  - **实时天气**: query_weather(city="城市名", forecast=False)
  - **未来预报**: query_weather(city="城市名", forecast=True, filter_indices="索引")
    * 📋 filter_indices 参数说明：
      - 用于控制前端格式化展示哪些天的数据
      - 格式：用逗号分隔的索引（从0开始）
      - 0=第1天, 1=第2天, 2=第3天, 3=第4天
    * 🎯 使用示例：
      - 用户说"未来3天" → filter_indices="0,1,2"
      - 用户说"明天和后天" → filter_indices="0,1"
      - 用户说"周三周四"（假设是第3、4天）→ filter_indices="2,3"
      - 用户说"后天" → filter_indices="1"
      - 用户说"除了明天的其他天" → filter_indices="1,2,3"
    * ⚠️ 重要规则：
      1. 在文本回答中，详细说明用户要求的天数的天气情况
      2. 同时传递 filter_indices，让前端卡片展示与你的文字说明一致
      3. 这样用户看到的文字和卡片就完全对应了

**决策原则**：
1. 编程、知识类问题 → 直接回答
2. 日程查询 → 使用 get_company_schedule
3. 画图请求 → 使用 generate_image
4. 天气查询 → 使用 query_weather
5. 不确定时 → 直接回答

请用中文回复，保持友好和专业。"""),
        ("human", "{input}"),
        MessagesPlaceholder(variable_name="agent_scratchpad"),
    ])
    
    # 创建 Agent
    agent = create_openai_tools_agent(llm, tools, prompt)
    # ⚠️ 关键修改：设置 return_intermediate_steps=True
    agent_executor = AgentExecutor(
        agent=agent, 
        tools=tools, 
        verbose=True,
        return_intermediate_steps=True  # 必须开启才能获取工具返回值！
    )
    
    return agent_executor


# 聊天接口
@app.route('/api/chat', methods=['POST'])
def chat():
    """
    聊天接口
    请求体: {"message": "查询技术部员工日程"}
    """
    try:
        data = request.get_json()
        user_message = data.get('message', '')
        
        if not user_message:
            return jsonify({"error": "消息不能为空"}), 400
        
        # 创建 Agent 并执行
        agent_executor = create_agent()
        result = agent_executor.invoke({"input": user_message})
        
        print(f"\n{'='*60}")
        print(f"📥 Agent 执行结果:")
        print(f"{'='*60}")
        
        # 提取结果
        response_text = result.get('output', '')
        print(f"📝 AI回复: {response_text[:100]}...")
        
        # 尝试从响应中提取 JSON 数据（日程、图片或天气）
        schedule_data = None
        image_data = None
        weather_data = None
        
        # 方法1：从中间步骤中提取工具返回值
        intermediate_steps = result.get('intermediate_steps', [])
        print(f"🔍 中间步骤数量: {len(intermediate_steps)}")
        
        for i, step in enumerate(intermediate_steps):
            print(f"\n--- 步骤 {i+1} ---")
            if len(step) >= 2:
                action, observation = step
                
                # 获取工具名称
                tool_name = getattr(action, 'tool', 'unknown')
                print(f"🛠️  工具名称: {tool_name}")
                print(f"📊 返回数据长度: {len(str(observation))}")
                
                # 检查是否调用了画图工具
                if tool_name == 'generate_image':
                    print(f"🎨 检测到画图工具调用")
                    try:
                        image_data = json.loads(observation)
                        print(f"✅ 成功提取图片数据!")
                        print(f"   - image_id: {image_data.get('image_id', 'N/A')}")
                        print(f"   - url: {image_data.get('url', 'N/A')}")
                        print(f"   - success: {image_data.get('success', False)}")
                    except Exception as e:
                        print(f"❌ 解析图片数据失败: {e}")
                        print(f"   原始数据: {observation[:200]}")
                
                # 检查是否调用了日程工具
                elif tool_name == 'get_company_schedule':
                    print(f"📋 检测到日程工具调用")
                    try:
                        schedule_data = json.loads(observation)
                        print(f"✅ 成功提取日程数据!")
                    except Exception as e:
                        print(f"❌ 解析日程数据失败: {e}")
                
                # 检查是否调用了天气工具
                elif tool_name == 'query_weather':
                    print(f"🌤️  检测到天气工具调用（MCP 协议）")
                    try:
                        weather_data = json.loads(observation)
                        print(f"✅ 成功提取天气数据!")
                        if weather_data.get('success'):
                            print(f"   - 城市: {weather_data.get('city')}")
                            print(f"   - 天气: {weather_data.get('weather')}")
                            print(f"   - 温度: {weather_data.get('temperature')}°C")
                    except Exception as e:
                        print(f"❌ 解析天气数据失败: {e}")
        
        # 方法2：从响应文本中提取 JSON（备用）
        if not image_data and not schedule_data and '{' in response_text:
            print(f"\n🔄 尝试从响应文本中提取JSON...")
            try:
                import re
                json_matches = list(re.finditer(r'\{[^{}]*\}', response_text))
                print(f"   找到 {len(json_matches)} 个可能的JSON对象")
                
                for json_match in json_matches:
                    try:
                        parsed_data = json.loads(json_match.group())
                        
                        # 检查是否为日程数据
                        if 'schedules' in parsed_data and not schedule_data:
                            schedule_data = parsed_data
                            print(f"✅ 从文本提取到日程数据")
                        # 检查是否为图片数据
                        elif ('image_id' in parsed_data or 'url' in parsed_data) and not image_data:
                            image_data = parsed_data
                            print(f"✅ 从文本提取到图片数据")
                    except:
                        continue
            except Exception as e:
                print(f"⚠️  文本提取失败: {e}")
        
        print(f"\n{'='*60}")
        print(f"📤 最终返回:")
        print(f"   - 日程数据: {'✅ 有' if schedule_data else '❌ 无'}")
        print(f"   - 图片数据: {'✅ 有' if image_data else '❌ 无'}")
        print(f"   - 天气数据: {'✅ 有' if weather_data else '❌ 无'}")
        if image_data:
            print(f"   - 图片URL: {image_data.get('url', 'N/A')}")
        if weather_data and weather_data.get('success'):
            print(f"   - 天气: {weather_data.get('city')} {weather_data.get('weather')} {weather_data.get('temperature')}°C")
        print(f"{'='*60}\n")
        
        return jsonify({
            "success": True,
            "message": response_text,
            "data": schedule_data,
            "image": image_data,
            "weather": weather_data
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


# 下载 Excel 接口
@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    """
    下载员工日程为 Excel 文件
    请求体: {"schedules": [...]}
    """
    try:
        data = request.get_json()
        schedules = data.get('schedules', [])
        
        if not schedules:
            return jsonify({"error": "没有可下载的数据"}), 400
        
        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工日程表"
        
        # 设置表头
        headers = ["部门", "员工姓名", "职位", "日期", "任务", "状态", "优先级"]
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加数据
        for schedule in schedules:
            ws.append([
                schedule.get("department", ""),
                schedule.get("employee_name", ""),
                schedule.get("position", ""),
                schedule.get("date", ""),
                schedule.get("task", ""),
                schedule.get("status", ""),
                schedule.get("priority", "")
            ])
        
        # 调整列宽
        column_widths = [15, 15, 18, 15, 30, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 保存到内存
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 生成文件名
        filename = f"员工日程表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# 图片访问接口
@app.route('/api/images/<filename>', methods=['GET'])
def serve_image(filename):
    """
    访问生成的图片
    """
    try:
        if IMAGE_CONFIG:
            return send_from_directory(IMAGES_DIR, filename)
        else:
            return jsonify({"error": "图片服务未配置"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 404


# 健康检查接口
@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查接口"""
    return jsonify({
        "status": "ok",
        "message": "服务运行正常"
    })


# 根路径欢迎页面
@app.route('/', methods=['GET'])
def index():
    """欢迎页面 - 显示使用说明"""
    html = """
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>公司员工日程管理 API</title>
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            body {
                font-family: 'Microsoft YaHei', Arial, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                display: flex;
                justify-content: center;
                align-items: center;
                padding: 20px;
            }
            .container {
                background: white;
                border-radius: 20px;
                padding: 40px;
                max-width: 800px;
                box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            }
            h1 {
                color: #667eea;
                margin-bottom: 20px;
                font-size: 32px;
            }
            .status {
                background: #dcfce7;
                color: #16a34a;
                padding: 15px;
                border-radius: 10px;
                margin-bottom: 30px;
                font-weight: 600;
            }
            .section {
                margin-bottom: 30px;
            }
            h2 {
                color: #333;
                margin-bottom: 15px;
                font-size: 20px;
                border-left: 4px solid #667eea;
                padding-left: 10px;
            }
            .instruction {
                background: #f8f9fa;
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 15px;
            }
            .instruction h3 {
                color: #667eea;
                margin-bottom: 10px;
                font-size: 16px;
            }
            .instruction p {
                color: #666;
                line-height: 1.6;
                margin-bottom: 10px;
            }
            .code {
                background: #2d2d2d;
                color: #f8f8f2;
                padding: 15px;
                border-radius: 8px;
                font-family: 'Consolas', monospace;
                margin: 10px 0;
                overflow-x: auto;
            }
            .button {
                display: inline-block;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 12px 30px;
                border-radius: 25px;
                text-decoration: none;
                font-weight: 600;
                margin-top: 10px;
                transition: transform 0.3s ease;
            }
            .button:hover {
                transform: translateY(-2px);
            }
            .api-list {
                list-style: none;
            }
            .api-list li {
                background: #f8f9fa;
                padding: 10px 15px;
                margin-bottom: 8px;
                border-radius: 8px;
                border-left: 3px solid #667eea;
            }
            .api-list code {
                color: #667eea;
                font-weight: 600;
            }
            .warning {
                background: #fef3c7;
                color: #d97706;
                padding: 15px;
                border-radius: 10px;
                margin-top: 20px;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🤖 公司员工日程管理 API</h1>
            
            <div class="status">
                ✅ 后端服务运行正常！
            </div>
            
            <div class="section">
                <h2>📖 如何使用聊天机器人？</h2>
                
                <div class="instruction">
                    <h3>方式 1：直接打开前端页面（推荐）</h3>
                    <p>1. 在项目文件夹中找到 <strong>index.html</strong> 文件</p>
                    <p>2. <strong>双击打开</strong>或右键选择"用浏览器打开"</p>
                    <p>3. 开始与 AI 助手对话！</p>
                </div>
                
                <div class="instruction">
                    <h3>方式 2：使用文件路径</h3>
                    <p>在浏览器地址栏输入：</p>
                    <div class="code">file:///你的路径/cursor-web-project/index.html</div>
                </div>
                
                <div class="instruction">
                    <h3>方式 3：使用 VSCode Live Server</h3>
                    <p>在 VSCode 中右键 index.html → Open with Live Server</p>
                </div>
            </div>
            
            <div class="section">
                <h2>🔌 可用的 API 接口</h2>
                <ul class="api-list">
                    <li><code>GET /api/health</code> - 健康检查</li>
                    <li><code>POST /api/employee-schedule</code> - 获取员工日程</li>
                    <li><code>POST /api/chat</code> - AI 聊天接口</li>
                    <li><code>POST /api/download-excel</code> - 下载 Excel</li>
                </ul>
            </div>
            
            <div class="section">
                <h2>🏢 支持的部门</h2>
                <p style="color: #666; line-height: 1.8;">
                    技术部 • 市场部 • 人事部 • 财务部
                </p>
            </div>
            
            <div class="warning">
                ⚠️ <strong>注意：</strong>这是后端 API 服务地址，不是聊天页面。<br>
                请打开 <strong>index.html</strong> 文件来使用聊天机器人。
            </div>
        </div>
    </body>
    </html>
    """
    return html


if __name__ == '__main__':
    print("=" * 60)
    print("🤖 公司员工日程管理聊天机器人启动中...")
    print("=" * 60)
    print("📋 可用的部门: 技术部、市场部、人事部、财务部")
    print("🌐 后端服务地址: http://localhost:5000")
    print("💡 前端页面: 请在浏览器打开 index.html")
    print("=" * 60)
    
    app.run(debug=True, host='0.0.0.0', port=5000)

